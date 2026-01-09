#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ConforME - Exportacao para Excel
================================

Script 03/03 da automacao de compliance de marketing.

Responsabilidades:
- Ler resultados JSON gerados pelo avaliacao_ia.py
- Gerar planilha Excel independente (por execucao)
- Atualizar planilha cumulativa (historico master)
- Incluir todas as colunas obrigatorias

Autor: ConforME Team
Data: Janeiro 2026
"""

import os
import sys
import json
import yaml
import logging
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# Importacao condicional do openpyxl (unico uso permitido: geracao de Excel)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


# ============================================================================
# CONFIGURACAO DE PATHS
# ============================================================================
BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "config" / "config.yaml"
LOGS_DIR = BASE_DIR / "logs"


# ============================================================================
# LOGGING
# ============================================================================
def setup_logging(config: Dict[str, Any]) -> logging.Logger:
    """
    Configura o sistema de logging.
    
    Args:
        config: Dicionario de configuracoes do YAML.
        
    Returns:
        Logger configurado.
    """
    log_config = config.get("logging", {})
    log_level = getattr(logging, log_config.get("level", "INFO").upper())
    log_format = log_config.get("format", "%(asctime)s | %(levelname)-8s | %(message)s")
    
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    
    handlers = [
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            LOGS_DIR / "exportacao.log",
            encoding="utf-8"
        )
    ]
    
    logging.basicConfig(
        level=log_level,
        format=log_format,
        handlers=handlers
    )
    
    return logging.getLogger("exportacao")


# ============================================================================
# CONFIGURACAO
# ============================================================================
def load_config() -> Dict[str, Any]:
    """
    Carrega configuracoes do arquivo config.yaml.
    
    Returns:
        Dicionario com configuracoes.
    """
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Arquivo de configuracao nao encontrado: {CONFIG_PATH}")
    
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    
    return config


# ============================================================================
# CARREGAMENTO DE RESULTADOS
# ============================================================================
def load_results(json_path: Path, logger: logging.Logger) -> Dict[str, Any]:
    """
    Carrega resultados do arquivo JSON.
    
    Args:
        json_path: Caminho do arquivo JSON.
        logger: Logger.
        
    Returns:
        Dicionario com resultados.
    """
    if not json_path.exists():
        raise FileNotFoundError(f"Arquivo de resultados nao encontrado: {json_path}")
    
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    logger.info(f"Resultados carregados: {json_path}")
    logger.info(f"Total de registros: {len(data.get('resultados', []))}")
    
    return data


# ============================================================================
# DETERMINACAO DO RESULTADO
# ============================================================================
def determinar_resultado(campos: Dict[str, str], status: str, erro: str) -> str:
    """
    Determina o resultado baseado nos campos extraidos.
    
    Args:
        campos: Campos extraidos da resposta da IA.
        status: Status do processamento.
        erro: Mensagem de erro, se houver.
        
    Returns:
        Resultado: Aprovado, Reprovado, Inconclusivo ou vazio.
    """
    # Se houve erro no processamento, retorna vazio
    if status == "erro" or erro:
        return ""
    
    # Verifica o resultado da IA
    resultado_ia = campos.get("RESULTADO", "").upper().strip()
    
    if resultado_ia in ["APROVADO", "APROVADA"]:
        return "Aprovado"
    elif resultado_ia in ["REPROVADO", "REPROVADA"]:
        return "Reprovado"
    elif resultado_ia in ["INCONCLUSIVO", "INCONCLUSIVA"]:
        return "Inconclusivo"
    
    # Se nao conseguiu determinar pelo campo RESULTADO, analisa violacoes
    violacoes = campos.get("VIOLACOES_ENCONTRADAS", "").strip()
    
    if violacoes and violacoes.lower() not in ["nenhuma", "nao", "n/a", "-", ""]:
        return "Reprovado"
    
    # Se tem avaliacao mas sem violacoes claras
    avaliacao = campos.get("AVALIACAO", "").strip()
    if avaliacao and not violacoes:
        return "Aprovado"
    
    return "Inconclusivo"


def resumir_avaliacao(campos: Dict[str, str], max_chars: int = 500) -> str:
    """
    Resume a avaliacao em ate 500 caracteres.
    
    Args:
        campos: Campos extraidos da resposta da IA.
        max_chars: Maximo de caracteres.
        
    Returns:
        Resumo da avaliacao.
    """
    avaliacao = campos.get("AVALIACAO", "").strip()
    
    if not avaliacao:
        # Tenta montar a partir de outros campos
        violacoes = campos.get("VIOLACOES_ENCONTRADAS", "").strip()
        justificativa = campos.get("JUSTIFICATIVA", "").strip()
        
        if violacoes and violacoes.lower() not in ["nenhuma", "nao", "n/a", "-"]:
            avaliacao = f"Violacoes: {violacoes}"
            if justificativa:
                avaliacao += f" | {justificativa}"
        elif justificativa:
            avaliacao = justificativa
    
    # Limita ao tamanho maximo
    if len(avaliacao) > max_chars:
        avaliacao = avaliacao[:max_chars-3] + "..."
    
    return avaliacao


# ============================================================================
# PREPARACAO DE DADOS
# ============================================================================
def prepare_excel_data(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Prepara dados dos resultados para formato tabular do Excel.
    
    Args:
        results: Dicionario com resultados da avaliacao.
        
    Returns:
        Lista de dicionarios prontos para Excel.
    """
    rows = []
    
    for item in results.get("resultados", []):
        campos = item.get("campos_extraidos", {})
        status = item.get("status", "")
        erro = item.get("erro", "") or ""
        
        # Determina resultado
        resultado = determinar_resultado(campos, status, erro)
        
        # Resume avaliacao
        avaliacao_resumida = resumir_avaliacao(campos)
        
        # Monta caminho completo
        caminho_completo = item.get("caminho", "") or ""
        if not caminho_completo:
            pasta_origem = item.get("pasta_origem", "")
            arquivo = item.get("arquivo", "")
            if pasta_origem and arquivo:
                caminho_completo = str(Path(pasta_origem) / arquivo)
        
        row = {
            "Data": datetime.fromisoformat(item.get("data_avaliacao", "")).strftime("%d/%m/%Y %H:%M")
                   if item.get("data_avaliacao") else "",
            "Nome do Arquivo": item.get("arquivo", ""),
            "Caminho Pasta": caminho_completo,
            "Hash SHA256": item.get("hash_sha256", ""),
            "Conteudo Identificado": campos.get("CONTEUDO_IDENTIFICADO", ""),
            "Violacoes Encontradas": campos.get("VIOLACOES_ENCONTRADAS", ""),
            "Avaliacao": avaliacao_resumida,
            "Resultado": resultado,
            "Justificativa": campos.get("JUSTIFICATIVA", ""),
            "Recomendacoes": campos.get("RECOMENDACOES", ""),
            "Status Processamento": status,
            "Erro": erro,
            "Parecer Final Humano": ""  # Sempre vazio inicialmente
        }
        
        rows.append(row)
    
    return rows


# ============================================================================
# ESTILOS DO EXCEL
# ============================================================================
def get_header_style():
    """Retorna estilo para cabecalho do Excel."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }


def get_result_fill(result: str) -> PatternFill:
    """
    Retorna cor de fundo baseada no resultado.
    
    Args:
        result: Resultado da avaliacao.
        
    Returns:
        PatternFill com cor correspondente.
    """
    result_upper = result.upper().strip() if result else ""
    
    colors = {
        "APROVADO": "C6EFCE",      # Verde claro
        "REPROVADO": "FFC7CE",     # Vermelho claro
        "INCONCLUSIVO": "FFEB9C"   # Amarelo claro
    }
    
    color = colors.get(result_upper, "FFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_cell_style():
    """Retorna estilo para celulas de dados."""
    return {
        "alignment": Alignment(vertical="top", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }


# ============================================================================
# GERACAO DE EXCEL
# ============================================================================
def create_excel_workbook(data: List[Dict[str, Any]], logger: logging.Logger) -> Workbook:
    """
    Cria workbook do Excel com os dados formatados.
    
    Args:
        data: Lista de dicionarios com dados.
        logger: Logger.
        
    Returns:
        Workbook pronto para salvar.
    """
    if not data:
        raise ValueError("Nenhum dado para exportar")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Compliance"
    
    # Colunas
    columns = list(data[0].keys())
    header_style = get_header_style()
    cell_style = get_cell_style()
    
    # Define larguras customizadas
    column_widths = {
        "Data": 18,
        "Nome do Arquivo": 30,
        "Caminho Pasta": 50,
        "Hash SHA256": 15,
        "Conteudo Identificado": 40,
        "Violacoes Encontradas": 35,
        "Avaliacao": 60,
        "Resultado": 15,
        "Justificativa": 40,
        "Recomendacoes": 40,
        "Status Processamento": 15,
        "Erro": 40,
        "Parecer Final Humano": 40
    }
    
    # Cabecalho
    for col_idx, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.alignment = header_style["alignment"]
        cell.border = header_style["border"]
        
        # Define largura
        width = column_widths.get(column_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Dados
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, column_name in enumerate(columns, 1):
            value = row_data.get(column_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_style["alignment"]
            cell.border = cell_style["border"]
            
            # Cor especial para coluna Resultado
            if column_name == "Resultado":
                cell.fill = get_result_fill(value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Congela cabecalho
    ws.freeze_panes = "A2"
    
    # Altura da linha do cabecalho
    ws.row_dimensions[1].height = 30
    
    logger.info(f"Workbook criado: {len(data)} linhas, {len(columns)} colunas")
    
    return wb


def save_independent_excel(
    wb: Workbook,
    config: Dict[str, Any],
    logger: logging.Logger
) -> Path:
    """
    Salva planilha Excel independente (por execucao).
    
    Args:
        wb: Workbook a salvar.
        config: Configuracoes.
        logger: Logger.
        
    Returns:
        Caminho do arquivo salvo.
    """
    output_dir = BASE_DIR / config["paths"]["output_folder"]
    output_dir.mkdir(parents=True, exist_ok=True)
    
    export_config = config.get("export", {})
    prefix = export_config.get("filename_prefix", "ResultadoConforme")
    date_format = export_config.get("date_format", "%d%m%Y")
    
    date_str = datetime.now().strftime(date_format)
    filename = f"{prefix}{date_str}.xlsx"
    filepath = output_dir / filename
    
    # Se ja existe, adiciona timestamp
    if filepath.exists():
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"{prefix}{date_str}_{timestamp}.xlsx"
        filepath = output_dir / filename
    
    wb.save(filepath)
    logger.info(f"Planilha independente salva: {filepath}")
    
    return filepath


def update_master_excel(
    data: List[Dict[str, Any]],
    config: Dict[str, Any],
    logger: logging.Logger
) -> Path:
    """
    Atualiza planilha cumulativa (historico master).
    
    Args:
        data: Dados a adicionar.
        config: Configuracoes.
        logger: Logger.
        
    Returns:
        Caminho do arquivo master.
    """
    output_dir = BASE_DIR / config["paths"]["output_folder"]
    output_dir.mkdir(parents=True, exist_ok=True)
    
    master_filename = config.get("export", {}).get("master_filename", "historico_master.xlsx")
    master_path = output_dir / master_filename
    
    if master_path.exists():
        # Abre arquivo existente
        wb = load_workbook(master_path)
        ws = wb.active
        
        # Encontra ultima linha
        last_row = ws.max_row
        
        # Adiciona novos dados
        columns = list(data[0].keys()) if data else []
        cell_style = get_cell_style()
        
        for row_data in data:
            last_row += 1
            for col_idx, column_name in enumerate(columns, 1):
                value = row_data.get(column_name, "")
                cell = ws.cell(row=last_row, column=col_idx, value=value)
                cell.alignment = cell_style["alignment"]
                cell.border = cell_style["border"]
                
                if column_name == "Resultado":
                    cell.fill = get_result_fill(value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"Adicionados {len(data)} registros ao historico master")
        
    else:
        # Cria novo arquivo master
        wb = create_excel_workbook(data, logger)
        logger.info("Criado novo arquivo historico master")
    
    wb.save(master_path)
    logger.info(f"Historico master atualizado: {master_path}")
    
    return master_path


# ============================================================================
# MAIN
# ============================================================================
def main() -> int:
    """
    Funcao principal do script de exportacao.
    
    Returns:
        Codigo de saida (0 = sucesso, 1 = erro).
    """
    print("\n" + "=" * 60)
    print("ConforME - Exportacao para Excel")
    print("=" * 60 + "\n")
    
    # Verifica dependencias
    if not EXCEL_SUPPORT:
        print("[ERRO] openpyxl nao instalado. Use: pip install openpyxl")
        return 1
    
    # Parse argumentos
    parser = argparse.ArgumentParser(description="Exportacao de resultados para Excel")
    parser.add_argument(
        "--input", "-i",
        type=str,
        required=True,
        help="Caminho do arquivo JSON com resultados da avaliacao"
    )
    args = parser.parse_args()
    
    try:
        # Carrega configuracoes
        config = load_config()
        logger = setup_logging(config)
        
        logger.info("Iniciando processo de exportacao")
        
        # Carrega resultados
        logger.info("-" * 40)
        logger.info("ETAPA 1: Carregamento de resultados")
        logger.info("-" * 40)
        
        json_path = Path(args.input)
        results = load_results(json_path, logger)
        
        # Prepara dados
        logger.info("-" * 40)
        logger.info("ETAPA 2: Preparacao dos dados")
        logger.info("-" * 40)
        
        data = prepare_excel_data(results)
        
        if not data:
            logger.warning("Nenhum resultado para exportar")
            return 0
        
        logger.info(f"Dados preparados: {len(data)} registros")
        
        # Gera Excel independente
        logger.info("-" * 40)
        logger.info("ETAPA 3: Geracao do Excel independente")
        logger.info("-" * 40)
        
        wb = create_excel_workbook(data, logger)
        independent_path = save_independent_excel(wb, config, logger)
        
        # Atualiza master
        logger.info("-" * 40)
        logger.info("ETAPA 4: Atualizacao do historico master")
        logger.info("-" * 40)
        
        master_path = update_master_excel(data, config, logger)
        
        # Resumo final
        # Conta resultados
        approved = sum(1 for d in data if d.get("Resultado", "").upper() == "APROVADO")
        rejected = sum(1 for d in data if d.get("Resultado", "").upper() == "REPROVADO")
        inconclusive = sum(1 for d in data if d.get("Resultado", "").upper() == "INCONCLUSIVO")
        errors = len(data) - approved - rejected - inconclusive
        
        logger.info("=" * 40)
        logger.info("EXPORTACAO CONCLUIDA")
        logger.info("=" * 40)
        logger.info(f"Total: {len(data)} registros")
        logger.info(f"  Aprovados: {approved}")
        logger.info(f"  Reprovados: {rejected}")
        logger.info(f"  Inconclusivos: {inconclusive}")
        logger.info(f"  Com erro: {errors}")
        
        print(f"\n[OK] Exportacao concluida!")
        print(f"   [EXCEL] Planilha independente: {independent_path}")
        print(f"   [MASTER] Historico master: {master_path}")
        print(f"\n   Resumo:")
        print(f"      Total: {len(data)} | Aprovados: {approved} | Reprovados: {rejected} | Inconclusivos: {inconclusive} | Erros: {errors}\n")
        
        return 0
        
    except FileNotFoundError as e:
        print(f"\n[ERRO] Arquivo nao encontrado: {e}")
        return 1
    except ValueError as e:
        print(f"\n[ERRO] Erro de dados: {e}")
        return 1
    except Exception as e:
        print(f"\n[ERRO] Erro inesperado: {e}")
        logging.exception("Erro fatal na execucao")
        return 1


if __name__ == "__main__":
    sys.exit(main())
